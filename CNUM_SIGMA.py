# -*- coding: utf-8 -*-

# =============================================================================
# =============================================================================
#                           AGENDA SIGMA
# =============================================================================
# =============================================================================


# CHARGEMENT DES BIBLIOTHEQUES
# ============================

import os
import re
import csv
from io import BytesIO
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow


# CONFIGURATION DU CODE : definition des chemins d'accès
# ============================

#Lit le fichier de configuration et renvoie un dictionnaire contenant les chemins et paramètres.
#Chaque ligne doit être au format : clé = valeur.
#Les lignes vides ou commençant par '#' sont ignorées.

def read_config(config_file="config.txt"):
    config = {}
    # On utilise le répertoire du script pour construire le chemin absolu du fichier config.txt
    if os.path.exists(config_file):
        with open(config_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, value = line.split("=", 1)
                    config[key.strip()] = value.strip()
    else:
            print(f"Fichier de configuration '{config_file}' non trouvé dans {script_dir}. Les chemins par défaut seront utilisés.")
    return config

# Lecture du fichier de configuration
config = read_config()


# =============================================================================
# PARTIE  1: SELECTION DU TABLEAU D'INTERET
# =============================================================================

# Définition des chemins et paramètres de la feuille d'intérêt via le fichier de configuration
file_path = config.get("excel_file_path")
sheet_name = config.get("sheet_name", "M1 2324")  # Nom de la feuille contenant l'emploi du temps ou les données d'intérêt

print("Début de l'exécution du script...")

# ------------------------------------------------------------------
# Chargement du fichier Excel original avec openpyxl pour récupérer
# les informations de formatage (couleurs, commentaires) et les cellules fusionnées
# ------------------------------------------------------------------
wb_orig = load_workbook(file_path)
ws_orig = wb_orig[sheet_name]
merged_cells = []
for merge_range in ws_orig.merged_cells.ranges:
    if (merge_range.min_col >= 5 and merge_range.max_col <= 15 and 
        merge_range.min_row >= 5 and merge_range.max_row <= 34):
        merged_cells.append(merge_range)

# ------------------------------------------------------------------
# Chargement des données dans un DataFrame avec pandas.
# - On ignore les 4 premières lignes (skiprows=4) pour se positionner sur la zone utile.
# - On sélectionne les colonnes E à O (usecols="E:O") pour extraire uniquement le tableau souhaité.
# - La ligne lue en première position (après skiprows) sert d'en-tête (header=0).
# - On limite le DataFrame aux 29 premières lignes de données.
# ------------------------------------------------------------------
df = pd.read_excel(file_path, sheet_name=sheet_name,
                   skiprows=4, usecols="E:O", header=0, engine="openpyxl")
df = df.iloc[:29]

# ------------------------------------------------------------------
# Récupération des couleurs de remplissage et des commentaires des cellules
# dans la zone d'intérêt du fichier original.
# La zone considérée ici est de la ligne 6 à 34 et des colonnes 5 à 15.
# Les indices sont adaptés pour correspondre à la numérotation 0-based utilisée par pandas.
# ------------------------------------------------------------------
cell_colors = {}
cell_comments = {}
for row_idx, row in enumerate(ws_orig.iter_rows(min_row=6, max_row=34, min_col=5, max_col=15), start=0):
    for col_idx, cell in enumerate(row, start=0):  # Les indices commencent à 0 pour aligner avec le DataFrame
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
            cell_colors[(row_idx, col_idx)] = cell.fill.fgColor.rgb
        if cell.comment:
            cell_comments[(row_idx, col_idx)] = cell.comment.text
           
# ------------------------------------------------------------------
# Écriture des données (sans formatage) dans un nouveau classeur Excel en mémoire.
# Le nouveau classeur contiendra l'en-tête en ligne 1 et les données à partir de la ligne 2.
# ------------------------------------------------------------------
output = BytesIO()
df.to_excel(output, sheet_name="M1 2324_modifie", index=False)
output.seek(0)

# ------------------------------------------------------------------
# Rechargement du nouveau classeur avec openpyxl afin d'appliquer
# le formatage (couleurs et commentaires) ainsi que la gestion des cellules fusionnées.
# ------------------------------------------------------------------
new_wb = load_workbook(output)
new_ws = new_wb["M1 2324_modifie"]

# Application des couleurs et commentaires récupérés sur chaque cellule correspondante
# dans le nouveau classeur (les données commencent en ligne 2).
for row_idx in range(len(df)):
    for col_idx in range(len(df.columns)):
        cell = new_ws.cell(row=row_idx + 2, column=col_idx + 1)  # Ajustement : ligne d'en-tête décalée d'une unité
        if (row_idx, col_idx) in cell_colors:
            color = cell_colors[(row_idx, col_idx)]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if (row_idx, col_idx) in cell_comments:
            cell.comment = openpyxl.comments.Comment(cell_comments[(row_idx, col_idx)], "Author")

# ------------------------------------------------------------------
# Gestion des cellules fusionnées :
#   Pour chaque plage de cellules fusionnées détectée dans l'original,
#   on copie la valeur, la couleur et le commentaire de la cellule en haut à gauche
#   vers toutes les cellules correspondantes dans la nouvelle feuille.
# Les indices sont recalculés car l'en-tête du nouveau classeur remplace la ligne 5 de l'original
# et la colonne E devient la première colonne.
# ------------------------------------------------------------------
for merge_range in merged_cells:
    new_min_row = merge_range.min_row - 4  # Décalage dû à l'en-tête
    new_max_row = merge_range.max_row - 4
    new_min_col = merge_range.min_col - 4  # Décalage des colonnes (colonne E devient colonne 1)
    new_max_col = merge_range.max_col - 4

    original_top_left = ws_orig.cell(row=merge_range.min_row, column=merge_range.min_col)
    top_left_value = original_top_left.value
    top_left_color = None
    top_left_comment = original_top_left.comment.text if original_top_left.comment else None
    
    if original_top_left.fill and original_top_left.fill.fgColor and original_top_left.fill.fgColor.rgb and original_top_left.fill.fgColor.rgb != "00000000":
        top_left_color = original_top_left.fill.fgColor.rgb

    for r in range(new_min_row, new_max_row + 1):
        for c in range(new_min_col, new_max_col + 1):
            new_cell = new_ws.cell(row=r, column=c)
            new_cell.value = top_left_value
            if top_left_color:
                new_cell.fill = PatternFill(start_color=top_left_color, end_color=top_left_color, fill_type="solid")
            if top_left_comment:
                new_cell.comment = openpyxl.comments.Comment(top_left_comment, "Author")
                

# =============================================================================
# PARTIE 2: CREATION DU FICHIER CSV
# =============================================================================

# Définition du chemin de sortie pour le fichier CSV
output_csv = os.path.join(os.path.dirname(file_path), "output.csv")

# Dictionnaire des horaires (clé = intitulé exact de la colonne, ex : "Lu Matin")
horaires = {
    "Lu Matin": ("08:30", "12:30"),
    "Lu Aprem": ("13:30", "17:30"),
    "Ma Matin": ("08:00", "12:00"),
    "Ma Aprem": ("13:30", "16:00"),
    "Me Matin": ("08:30", "12:30"),
    "Me Aprem": ("13:30", "17:30"),
    "Je Matin": ("08:30", "12:30"),
    "Je Aprem": ("13:30", "17:30"),
    "Ve Matin": ("08:30", "12:30"),
    "Ve Aprem": ("13:30", "17:30"),
}

# Dictionnaire des points de séparation (mid-times) pour scinder un créneau en deux
mid_times = {
    "Lu Matin": "10:30",
    "Lu Aprem": "15:30",
    "Ma Matin": "10:00",
    "Ma Aprem": "15:30",
    "Me Matin": "10:30",
    "Me Aprem": "15:30",
    "Je Matin": "10:30",
    "Je Aprem": "15:30",
    "Ve Matin": "10:30",
    "Ve Aprem": "15:30",
}

# Dictionnaire de correspondance couleurs -> lieux
#Les codes couleurs sont ceux trouvés sur Excel ou Libreoffice (correspondant au Hex_# dans couleurs personnalisées)
color_to_location = {
    "F8CBAD": "Salle UT2J sans ordi",
    "CCFFCC": "Salle ENSAT sans ordi",
    "99CCFF": "1003-Langue",
    "FF9933": "UT2J GS027",
    "FFCC66": "UT2J GS021",
    "E2F0D9": "703 (projet) ou alternance (entreprise)",
    "FAFA9E": "UT2J GS027",
    "F5BCE9": "UT2JGS025"
}

# Pour convertir un libellé de mois en nombre
months_fr = {
    "jan": 1, "janv": 1, "janv.": 1, "févr": 2, "fev": 2, "fev.": 2, "févr.": 2,
    "mars": 3, "mars.": 3, "avr": 4, "avr.": 4, "avril": 4,
    "mai": 5, "mai.": 5, "juin": 6, "juin.": 6,
    "juil": 7, "juil.": 7, "juillet": 7,
    "août": 8, "aout": 8, "aout.": 8, "août.": 8,
    "sept": 9, "sept.": 9, "septembre": 9,
    "oct": 10, "oct.": 10, "octobre": 10,
    "nov": 11, "nov.": 11, "novembre": 11,
    "dec": 12, "dec.": 12, "déc": 12, "déc.": 12, "décembre": 12
}

# Pour faire correspondre "Lu" -> 0 (lundi), "Ma" -> 1 (mardi), etc.
day_offsets = {
    "Lu": 0,
    "Ma": 1,
    "Me": 2,
    "Je": 3,
    "Ve": 4
}

# ---------------------------------------------------------------------------
# Fonction pour scinder le contenu d'une cellule si elle contient "/"
# et générer 1 ou 2 événements selon les cas
# ---------------------------------------------------------------------------
def split_subject_into_events(subject, date_str, halfday_label, location, description):
    """Retourne une liste d'événements [Subject, Date, Start, End, Location, Description]."""
    events = []
    start_time, end_time = horaires.get(halfday_label, ("", ""))
    mid_time = mid_times.get(halfday_label, "")

    if "/" in subject:
        parts = [p.strip() for p in subject.split("/")]
        if len(parts) == 2:
            first_part, second_part = parts
            if first_part in ["---", "X"]:
                if second_part not in ["---", "X"]:
                    events.append([second_part, date_str, mid_time, end_time, location, description])
            else:
                events.append([first_part, date_str, start_time, mid_time, location, description])
                if second_part not in ["---", "X"]:
                    events.append([second_part, date_str, mid_time, end_time, location, description])
        else:
            events.append([subject, date_str, start_time, end_time, location, description])
    else:
        events.append([subject, date_str, start_time, end_time, location, description])
    return events

# ---------------------------------------------------------------------------
# Lecture du nouveau classeur modifié en mémoire 
# ---------------------------------------------------------------------------
ws = new_ws  # Par exemple, on utilise la feuille "M1 2324_modifie". Cette feuille n'apparait pas mais c'est sur celle ci qu'on travaille esnuite.

# Lecture de l'entête (ligne 1)
headers = [cell.value for cell in ws[1] if cell.value is not None]
if len(headers) < 2:
    print("En-têtes insuffisantes dans la première ligne du fichier Excel.")
    exit(1)
halfday_headers = headers[1:]  # on enlève la 1re colonne (date)

# ---------------------------------------------------------------------------
# Ouverture du fichier CSV en écriture
# ---------------------------------------------------------------------------
with open(output_csv, "w", newline="", encoding="utf-8") as csvfile:
    writer = csv.writer(csvfile, delimiter=';')
    writer.writerow(["Subject", "Date", "Start Time", "End Time", "Location", "Description"])

    # Parcours des lignes à partir de la 2e
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        week_cell = row[0]  # cellule de la colonne A
        if not week_cell or not week_cell.value:
            continue

        week_info = str(week_cell.value).strip().lower()
        # On cherche un motif du type "11-15 sept 23" ou "18-22 sept. 2023"
        match = re.search(r"(\d+)\s*-\s*(\d+)\s+([a-zA-Zéû\.]+)\s+(\d+)", week_info)
        if not match:
            continue

        try:
            day_start = int(match.group(1))
            day_end = int(match.group(2))
            month_str = match.group(3).replace('.', '')
            year_str = match.group(4)

            # Conversion mois/année
            month = months_fr.get(month_str, None)
            if not month:
                continue

            if len(year_str) == 2:
                year = 2000 + int(year_str)
            else:
                year = int(year_str)

            # Si day_end < day_start => on suppose que day_start est le mois précédent
            # Exemple : "30-03 nov 23" => 30 (oct), 3 (nov)
            if day_end < day_start:
                new_month = month - 1
                new_year = year
                if new_month < 1:  # Si on est avant janvier => année précédente
                    new_month = 12
                    new_year -= 1

                # day_start est dans le "nouveau" mois
                monday_date = datetime(new_year, new_month, day_start)
            else:
                # Sinon, c'est le même mois
                monday_date = datetime(year, month, day_start)

        except Exception as e:
            print(f"Erreur de parsing sur '{week_info}' : {e}")
            continue


        # Parcours des colonnes B.. (les demi-journées)
        for col_index, cell in enumerate(row[1:], start=1):
            if col_index - 1 < len(halfday_headers):
                halfday_label = halfday_headers[col_index - 1]
            else:
                continue

            if (cell.value is None) and (cell.comment is None):
                continue

            day_abbr = halfday_label.split()[0]  # "Lu", "Ma", "Me", etc.
            offset = day_offsets.get(day_abbr, None)
            if offset is None:
                continue

            # Date de l'événement = lundi + offset
            event_date = monday_date + timedelta(days=offset)
            date_str = event_date.strftime("%Y-%m-%d")
            subject = str(cell.value).strip() if cell.value else ""
            description = cell.comment.text.strip() if cell.comment else ""
            # Récupération de la couleur pour la salle
            location = ""
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = cell.fill.fgColor.rgb
                if rgb.startswith("FF") and len(rgb) == 8:
                    color_code = rgb[2:]
                else:
                    color_code = rgb
                location = color_to_location.get(color_code, "")

            events_to_write = split_subject_into_events(subject, date_str, halfday_label, location, description)
            for ev in events_to_write:
                writer.writerow(ev)

print(f"✅ Fichier CSV généré : {output_csv}")


# ---------------------------------------------------------------------------
# PARTIE 3 : CREATION DE L'AGENDA GOOGLE
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Authentification : Authentifie l'utilisateur et renvoie un service Google Calendar.
# Si erreur à ce niveau, il faut supprimer le fichier token.json du dossier.
# Cela va forcer le script à vous demander de vous réauthentifier avec votre adresse mail.
# ---------------------------------------------------------------------------

# Autorisations Google Calendar
SCOPES = ["https://www.googleapis.com/auth/calendar"]

def authenticate_google(token_path="token.json", credentials_path="credentials.json"):
    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as token:
            token.write(creds.to_json())

    return build("calendar", "v3", credentials=creds)

#Convertit une date et une heure en objet datetime.
def convert_to_datetime(date_str, time_str):
    datetime_str = f"{date_str} {time_str}"
    return datetime.strptime(datetime_str, "%Y-%m-%d %H:%M")  

#Génère un identifiant unique. 
def sanitize_csv_id(input_str):
    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789_")
    result = input_str.replace(" ", "_").lower()
    return ''.join(c for c in result if c in allowed)

#Charge tous les événements existants avec leur csv_id. But = limiter le nombre de requêtes afin qu'elles puissent toutes être prises en compte.
def fetch_existing_events(service):
    events_result = service.events().list(
        calendarId='primary',
        maxResults=1000,  # Adapter selon le nombre d'événements attendus
        singleEvents=True
    ).execute()

    existing_events = {}
    for event in events_result.get('items', []):
        csv_id = event.get('extendedProperties', {}).get('private', {}).get('csv_id')
        if csv_id:
            existing_events[csv_id] = event  # Stocke l'événement existant par son ID unique

    return existing_events

# Associer des couleurs aux salles dans Google Calendar. Il y en a 11 disponibles dans Google Calendar. 
def sync_events(service, df):
    room_colors = {  
        "Salle UT2J sans ordi": 11,  
        "UT2J GS027": 6,  
        "UT2J GS021": 5,  
        "1003-Langue": 7,  
        "Salle ENSAT sans ordi": 10,
        "703 (projet) ou alternance (entreprise)": 2,
        "UT2J GS028": 4,
    }

    existing_events = fetch_existing_events(service)  # Récupération en une seule requête
    events_to_create = []
    events_to_update = []
    events_to_delete = []

    for _, row in df.iterrows():
        raw_id = f"{row['Date']}_{row['Start Time']}_{row['Subject']}"
        csv_id = sanitize_csv_id(raw_id)
        
        start_datetime = convert_to_datetime(row['Date'], row['Start Time'])
        end_datetime = convert_to_datetime(row['Date'], row['End Time'])

        # Récupère la couleur associée à la salle, ou utilise une couleur par défaut
        location = row['Location']
        color_id = room_colors.get(location, 5)  # Par défaut, utilise la couleur 5 si la salle n'est pas dans le dictionnaire

        event_body = {
            'summary': row['Subject'],
            'location': location,
            'description': row['Description'],
            'start': {'dateTime': start_datetime.isoformat(), 'timeZone': 'Europe/Paris'},
            'end': {'dateTime': end_datetime.isoformat(), 'timeZone': 'Europe/Paris'},
            'extendedProperties': {'private': {'csv_id': csv_id}},
            'colorId': color_id  # Ajoute l'identifiant de couleur
        }

        if csv_id in existing_events:
            existing_event = existing_events[csv_id]
            differences = []
            if existing_event.get('summary', '') != row['Subject']:
                differences.append('summary')
            if existing_event.get('location', '') != location:
                differences.append('location')
            if existing_event.get('description', '') != row['Description']:
                differences.append('description')
            if existing_event.get('start', {}).get('dateTime', '') != start_datetime.isoformat():
                differences.append('start time')
            if existing_event.get('end', {}).get('dateTime', '') != end_datetime.isoformat():
                differences.append('end time')

            if differences:
                event_body['id'] = existing_event['id']  # Nécessaire pour la mise à jour en batch
                events_to_update.append(event_body)
        else:
            events_to_create.append(event_body)

    # Suppression des événements obsolètes afin qu'il n'y ai pas de superposition d'évènements.
    # Si l'évènement existe déjà, il est conservé. S'il y a la moindre modification, il est supprimé puis un nouvel évènement est crée avec les mises à jour 
    existing_csv_ids = set(existing_events.keys())
    new_csv_ids = set(sanitize_csv_id(f"{row['Date']}_{row['Start Time']}_{row['Subject']}") for _, row in df.iterrows())
    obsolete_ids = existing_csv_ids - new_csv_ids
    for obsolete_id in obsolete_ids:
        events_to_delete.append(existing_events[obsolete_id]['id'])

    print(f"Création : {len(events_to_create)} | Mise à jour : {len(events_to_update)} | Suppression : {len(events_to_delete)}")

    batch = service.new_batch_http_request()
    for event in events_to_create:
        batch.add(service.events().insert(calendarId='primary', body=event))
    for event in events_to_update:
        batch.add(service.events().update(calendarId='primary', eventId=event['id'], body=event))
    for event_id in events_to_delete:
        batch.add(service.events().delete(calendarId='primary', eventId=event_id))
    
    batch.execute()  # Envoie toutes les requêtes en une seule fois
    #--> permet d'éviter la saturation du système et de ne pas atteindre la limite de l'API Google Calendar

def main():
    # Lecture du fichier de configuration contenant tous les chemins d'accès
    config = read_config()
    token_path = config.get("token_path", "token.json")
    credentials_path = config.get("credentials_path", "credentials.json")
    output_csv = config.get("output_csv")
    
    service = authenticate_google(token_path, credentials_path)
    
    df = pd.read_csv(output_csv, sep=';')
    df.columns = df.columns.str.strip()  # Nettoie les colonnes
    
    sync_events(service, df)

if __name__ == "__main__":
    main()


