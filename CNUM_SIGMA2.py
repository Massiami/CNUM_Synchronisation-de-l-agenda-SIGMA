# -*- coding: utf-8 -*-

# =============================================================================
# =============================================================================
#                           AGENDA SIGMA
# =============================================================================
# =============================================================================
"""
Script complet pour la gestion de l'agenda Sigma :
- Surveillance des modifications dans un fichier Excel (journal de modifications)
- Génération d'un CSV (output.csv) avec les événements extraits (en tenant compte du formatage, couleurs, commentaires et cellules fusionnées)
- Synchronisation avec Google Calendar (création, mise à jour et suppression en batch)
"""

# CHARGEMENT DES BIBLIOTHEQUES
# ============================
import csv
import os
import time
import re
import threading
from io import BytesIO
from datetime import timedelta, datetime as dt, datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow

# =============================================================================
# PARTIE 1 : CONFIGURATION
# =============================================================================

def read_config(config_file="config.txt"):
    """
    Lit le fichier de configuration et retourne un dictionnaire contenant
    les chemins et paramètres. Chaque ligne doit être au format "clé = valeur".
    Les lignes vides ou commençant par '#' sont ignorées.
    """
    config = {}
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, config_file)
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, value = line.split("=", 1)
                    config[key.strip()] = value.strip()
    else:
        print(f"Fichier de configuration '{config_file}' non trouvé dans {script_dir}.")
    return config

# Lecture de la configuration
config = read_config()
print("📂 Chemins et paramètres récupérés depuis config.txt :")
for key, value in config.items():
    print(f"{key} = {value}")

FILE_PATH = config.get("excel_file_path")          # Fichier Excel d'origine
SHEET_NAME = config.get("sheet_name", "M1 2324")      # Nom de la feuille à traiter
CSV_MODIFICATIONS = config.get("modifications_csv", os.path.join(os.path.dirname(FILE_PATH), "journal_modifications.csv"))
OUTPUT_CSV = config.get("output_csv", os.path.join(os.path.dirname(FILE_PATH), "output.csv"))
TOKEN_PATH = config.get("token_path", "token.json")
CREDENTIALS_PATH = config.get("credentials_path", "credentials.json")

# =============================================================================
# PARTIE 2 : SURVEILLANCE DES MODIFICATIONS EXCEL
# =============================================================================

# Paramètres de surveillance
MIN_ROW = 6    # Première ligne de données
MAX_ROW = 34   # Dernière ligne de données
MIN_COL = 5    # Colonne E
MAX_COL = 15   # Colonne O
INTERVALLE_MODIF = 30  # Intervalle de vérification en secondes

# Correspondance des couleurs aux lieux pour la surveillance
COLOR_TO_LOCATION = {
    "F8CBAD": "Salle UT2J sans ordi",
    "CCFFCC": "Salle ENSAT sans ordi",
    "99CCFF": "1003-Langue",
    "FF9933": "UT2J GS027",
    "FFCC66": "UT2J GS021",
    "E2F0D9": "703 (projet) ou alternance (entreprise)",
    "FAFA9E": "UT2J GS027",
    "F5BCE9": "UT2JGS025",
}

def charger_etat_excel():
    """
    Charge l'état actuel des cellules du fichier Excel en tenant compte
    des cellules fusionnées sur la feuille spécifiée.
    Retourne un dictionnaire indexé par (feuille, coordonnée).
    """
    wb = load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    etat = {}
    merged_ranges = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        coord_fusion = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_ranges[f"{get_column_letter(col)}{row}"] = coord_fusion
    for row in ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL):
        for cell in row:
            coord = merged_ranges.get(cell.coordinate, cell.coordinate)
            couleur_hex = ""
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                couleur_hex = cell.fill.start_color.rgb[-6:]
            lieu = COLOR_TO_LOCATION.get(couleur_hex, "")
            valeur = cell.value
            commentaire = cell.comment.text if cell.comment else ""
            etat[(SHEET_NAME, coord)] = {
                "valeur": valeur,
                "lieu": lieu,
                "commentaire": commentaire
            }
    wb.close()
    return etat

def format_cell_data(cell_data):
    """Formate les données d'une cellule pour affichage dans le CSV."""
    parts = []
    valeur = cell_data.get("valeur")
    if valeur is None:
        valeur = ""
    parts.append(f"valeur: {valeur}")
    if cell_data.get("lieu"):
        parts.append(f"lieu: {cell_data.get('lieu')}")
    if cell_data.get("commentaire"):
        parts.append(f"commentaire: {cell_data.get('commentaire')}")
    return " | ".join(parts)

def comparer_etats(etat_precedent, etat_actuel):
    """
    Compare deux états du fichier Excel et retourne les modifications détectées.
    """
    modifications = []
    for cle, nouvelle_donnee in etat_actuel.items():
        ancienne_donnee = etat_precedent.get(cle, {"valeur": None, "lieu": "", "commentaire": ""})
        if ancienne_donnee != nouvelle_donnee:
            modifications.append({
                "date": dt.now().strftime("%Y-%m-%d"),
                "heure": dt.now().strftime("%H:%M:%S"),
                "cellule": cle[1],
                "ancienne_donnee": format_cell_data(ancienne_donnee),
                "nouvelle_donnee": format_cell_data(nouvelle_donnee)
            })
    return modifications

def enregistrer_modifications(modifications):
    """
    Enregistre les modifications détectées dans un fichier CSV.
    """
    fichier_existe = os.path.exists(CSV_MODIFICATIONS)
    with open(CSV_MODIFICATIONS, mode="a", newline="", encoding="utf-8") as file:
        writer = csv.writer(file, quoting=csv.QUOTE_ALL)
        if not fichier_existe:
            writer.writerow(["Date", "Heure", "Cellule", "Ancienne Donnée", "Nouvelle Donnée"])
        for modif in modifications:
            writer.writerow([
                modif["date"],
                modif["heure"],
                modif["cellule"],
                modif["ancienne_donnee"],
                modif["nouvelle_donnee"]
            ])
    with open(CSV_MODIFICATIONS, mode="r", encoding="utf-8") as file:
        data = file.readlines()
    print(f"[Modifications] CSV rechargé ({len(data)} lignes).")

def surveiller_excel():
    """
    Boucle infinie de surveillance des modifications dans le fichier Excel.
    """
    historique_modifications = []
    etat_precedent = charger_etat_excel()
    print("🔄 Surveillance des modifications Excel lancée sur la feuille", SHEET_NAME)
    while True:
        try:
            time.sleep(INTERVALLE_MODIF)
            etat_actuel = charger_etat_excel()
            modifications = comparer_etats(etat_precedent, etat_actuel)
            nouvelles_modifications = [modif for modif in modifications if modif not in historique_modifications]
            if nouvelles_modifications:
                enregistrer_modifications(nouvelles_modifications)
                print(f"✅ {len(nouvelles_modifications)} modification(s) enregistrée(s) à {dt.now().strftime('%Y-%m-%d %H:%M:%S')}")
                historique_modifications.extend(nouvelles_modifications)
                historique_modifications = historique_modifications[-100:]
            etat_precedent = etat_actuel
        except Exception as e:
            print(f"⚠️ Erreur de surveillance : {e}")
            time.sleep(INTERVALLE_MODIF)

# =============================================================================
# PARTIE 3 : GÉNÉRATION DU CSV POUR L'AGENDA ET EXTRACTION DES ÉVÉNEMENTS
# =============================================================================

# Dictionnaires de correspondance pour les horaires, couleurs et dates.
horaires = {
    "Lu Matin": ("08:30", "12:30"),
    "Lu Aprem": ("13:30", "17:30"),
    "Ma Matin": ("08:00", "12:00"),
    "Ma Aprem": ("13:30", "16:00"),
    "Me Matin": ("08:30", "12:30"),
    "Me Aprem": ("13:30", "17:30"),
    "Je Matin": ("08:30", "12:30"),
    "Je Aprem": ("13:30", "17:30"),
    "Ve Matin": ("08:30", "12:30"),
    "Ve Aprem": ("13:30", "17:30"),
}
mid_times = {
    "Lu Matin": "10:30",
    "Lu Aprem": "15:30",
    "Ma Matin": "10:00",
    "Ma Aprem": "15:30",
    "Me Matin": "10:30",
    "Me Aprem": "15:30",
    "Je Matin": "10:30",
    "Je Aprem": "15:30",
    "Ve Matin": "10:30",
    "Ve Aprem": "15:30",
}
color_to_location = {
    "F8CBAD": "Salle UT2J sans ordi",
    "CCFFCC": "Salle ENSAT sans ordi",
    "99CCFF": "1003-Langue",
    "FF9933": "UT2J GS027",
    "FFCC66": "UT2J GS021",
    "E2F0D9": "703 (projet) ou alternance (entreprise)",
    "FAFA9E": "UT2J GS027",
    "F5BCE9": "UT2JGS025"
}
months_fr = {
    "jan": 1, "janv": 1, "janv.": 1,
    "févr": 2, "fev": 2, "fev.": 2, "févr.": 2,
    "mars": 3, "mars.": 3,
    "avr": 4, "avr.": 4, "avril": 4,
    "mai": 5, "mai.": 5,
    "juin": 6, "juin.": 6,
    "juil": 7, "juil.": 7, "juillet": 7,
    "août": 8, "aout": 8, "aout.": 8, "août.": 8,
    "sept": 9, "sept.": 9, "septembre": 9,
    "oct": 10, "oct.": 10, "octobre": 10,
    "nov": 11, "nov.": 11, "novembre": 11,
    "dec": 12, "dec.": 12, "déc": 12, "déc.": 12, "décembre": 12
}
day_offsets = {"Lu": 0, "Ma": 1, "Me": 2, "Je": 3, "Ve": 4}

def split_subject_into_events(subject, date_str, halfday_label, location, description):
    """
    Scinde le contenu d'une cellule en un ou deux événements selon la présence de '/'.
    Retourne une liste d'événements sous la forme :
    [Subject, Date, Start Time, End Time, Location, Description]
    """
    events = []
    start_time, end_time = horaires.get(halfday_label, ("", ""))
    mid_time = mid_times.get(halfday_label, "")
    if "/" in subject:
        parts = [p.strip() for p in subject.split("/")]
        if len(parts) == 2:
            first_part, second_part = parts
            if first_part in ["---", "X"]:
                if second_part not in ["---", "X"]:
                    events.append([second_part, date_str, mid_time, end_time, location, description])
            else:
                events.append([first_part, date_str, start_time, mid_time, location, description])
                if second_part not in ["---", "X"]:
                    events.append([second_part, date_str, mid_time, end_time, location, description])
        else:
            events.append([subject, date_str, start_time, end_time, location, description])
    else:
        events.append([subject, date_str, start_time, end_time, location, description])
    return events

def process_agenda():
    """
    Traite la feuille Excel pour générer le CSV (OUTPUT_CSV) contenant les événements.
    Le script récupère le formatage (couleurs, commentaires, cellules fusionnées),
    extrait les événements en fonction des en-têtes et synchronise ensuite avec Google Calendar.
    """
    events_global = []
    wb_orig = load_workbook(FILE_PATH)
    try:
        ws_orig = wb_orig[SHEET_NAME]
    except KeyError:
        print(f"Feuille '{SHEET_NAME}' introuvable dans le fichier Excel.")
        wb_orig.close()
        return

    # Récupération des cellules fusionnées dans la zone d'intérêt
    merged_cells = []
    for merge_range in ws_orig.merged_cells.ranges:
        if (merge_range.min_col >= 5 and merge_range.max_col <= 15 and 
            merge_range.min_row >= 5 and merge_range.max_row <= 34):
            merged_cells.append(merge_range)
    
    # Chargement du tableau dans un DataFrame avec pandas
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME,
                       skiprows=4, usecols="E:O", header=0, engine="openpyxl")
    df = df.iloc[:29]
    
    # Récupération des couleurs et commentaires dans la zone d'intérêt
    cell_colors = {}
    cell_comments = {}
    for row_idx, row in enumerate(ws_orig.iter_rows(min_row=6, max_row=34, min_col=5, max_col=15), start=0):
        for col_idx, cell in enumerate(row, start=0):
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                cell_colors[(row_idx, col_idx)] = cell.fill.fgColor.rgb
            if cell.comment:
                cell_comments[(row_idx, col_idx)] = cell.comment.text

    # Écriture des données (sans formatage) dans un classeur Excel en mémoire
    output = BytesIO()
    df.to_excel(output, sheet_name=SHEET_NAME + "_modifie", index=False)
    output.seek(0)
    new_wb = load_workbook(output)
    new_ws = new_wb[SHEET_NAME + "_modifie"]

    # Application des couleurs et commentaires sur la nouvelle feuille
    for row_idx in range(len(df)):
        for col_idx in range(len(df.columns)):
            cell = new_ws.cell(row=row_idx + 2, column=col_idx + 1)
            if (row_idx, col_idx) in cell_colors:
                color = cell_colors[(row_idx, col_idx)]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if (row_idx, col_idx) in cell_comments:
                cell.comment = openpyxl.comments.Comment(cell_comments[(row_idx, col_idx)], "Author")

    # Gestion des cellules fusionnées
    for merge_range in merged_cells:
        new_min_row = merge_range.min_row - 4
        new_max_row = merge_range.max_row - 4
        new_min_col = merge_range.min_col - 4
        new_max_col = merge_range.max_col - 4
        original_top_left = ws_orig.cell(row=merge_range.min_row, column=merge_range.min_col)
        top_left_value = original_top_left.value
        top_left_color = None
        top_left_comment = original_top_left.comment.text if original_top_left.comment else None
        if (original_top_left.fill and original_top_left.fill.fgColor and 
            original_top_left.fill.fgColor.rgb and original_top_left.fill.fgColor.rgb != "00000000"):
            top_left_color = original_top_left.fill.fgColor.rgb
        for r in range(new_min_row, new_max_row + 1):
            for c in range(new_min_col, new_max_col + 1):
                new_cell = new_ws.cell(row=r, column=c)
                new_cell.value = top_left_value
                if top_left_color:
                    new_cell.fill = PatternFill(start_color=top_left_color, end_color=top_left_color, fill_type="solid")
                if top_left_comment:
                    new_cell.comment = openpyxl.comments.Comment(top_left_comment, "Author")
    
    # Récupération des en-têtes pour les demi-journées (la première colonne correspond à l'information de semaine)
    headers = [cell.value for cell in new_ws[1] if cell.value is not None]
    if len(headers) < 2:
        print(f"En-têtes insuffisantes dans la feuille {SHEET_NAME}.")
        wb_orig.close()
        return
    halfday_headers = headers[1:]
    
    # Parcours des lignes pour extraire les événements
    for row_idx, row in enumerate(new_ws.iter_rows(min_row=2), start=2):
        week_cell = row[0]
        if not week_cell or not week_cell.value:
            continue
        week_info = str(week_cell.value).strip().lower()
        match = re.search(r"(\d+)\s*-\s*(\d+)\s+([a-zA-Zéû\.]+)\s+(\d+)", week_info)
        if not match:
            continue
        try:
            day_start = int(match.group(1))
            day_end = int(match.group(2))
            month_str = match.group(3).replace('.', '')
            year_str = match.group(4)
            month = months_fr.get(month_str, None)
            if not month:
                continue
            year = int(year_str) if len(year_str) > 2 else 2000 + int(year_str)
            if day_end < day_start:
                new_month = month - 1
                new_year = year
                if new_month < 1:
                    new_month = 12
                    new_year -= 1
                monday_date = dt(new_year, new_month, day_start)
            else:
                monday_date = dt(year, month, day_start)
        except Exception as e:
            print(f"Erreur de parsing sur '{week_info}' dans la feuille {SHEET_NAME} : {e}")
            continue

        for col_index, cell in enumerate(row[1:], start=1):
            if col_index - 1 < len(halfday_headers):
                halfday_label = halfday_headers[col_index - 1]
            else:
                continue
            if (cell.value is None) and (cell.comment is None):
                continue
            day_abbr = halfday_label.split()[0]
            offset = day_offsets.get(day_abbr, None)
            if offset is None:
                continue
            event_date = monday_date + timedelta(days=offset)
            date_str = event_date.strftime("%Y-%m-%d")
            subject = str(cell.value).strip() if cell.value else ""
            description = cell.comment.text.strip() if cell.comment else ""
            # Récupération de la couleur pour déterminer la salle
            location = ""
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = cell.fill.fgColor.rgb
                if rgb.startswith("FF") and len(rgb) == 8:
                    color_code = rgb[2:]
                else:
                    color_code = rgb
                location = color_to_location.get(color_code, "")
            events = split_subject_into_events(subject, date_str, halfday_label, location, description)
            events_global.extend(events)
    wb_orig.close()
    
    # Écriture du CSV final pour l'agenda
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile, delimiter=';')
        writer.writerow(["Subject", "Date", "Start Time", "End Time", "Location", "Description"])
        for ev in events_global:
            writer.writerow(ev)
    print(f"✅ Fichier CSV généré pour l'agenda : {OUTPUT_CSV}")
    
    # Synchronisation avec Google Calendar
    service = authenticate_google(TOKEN_PATH, CREDENTIALS_PATH)
    df_agenda = pd.read_csv(OUTPUT_CSV, sep=';')
    df_agenda.columns = df_agenda.columns.str.strip()
    sync_events(service, df_agenda)

# =============================================================================
# PARTIE 4 : SYNCHRONISATION AVEC GOOGLE CALENDAR
# =============================================================================

SCOPES = ["https://www.googleapis.com/auth/calendar"]

def authenticate_google(token_path=TOKEN_PATH, credentials_path=CREDENTIALS_PATH):
    """
    Authentifie l'utilisateur auprès de Google Calendar et retourne le service.
    Si le token est expiré ou absent, le flux d'authentification est lancé.
    """
    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as token:
            token.write(creds.to_json())
    return build("calendar", "v3", credentials=creds)

def convert_to_datetime(date_str, time_str):
    """
    Convertit une date (format 'YYYY-MM-DD') et une heure (format 'HH:MM') en objet datetime.
    """
    datetime_str = f"{date_str} {time_str}"
    return dt.strptime(datetime_str, "%Y-%m-%d %H:%M")

def sanitize_csv_id(input_str):
    """
    Génère un identifiant unique en conservant uniquement les caractères alphanumériques en minuscules et le caractère '_'.
    On combine plusieurs champs (Date, Start Time, Subject, Location) pour obtenir un identifiant stable.
    """
    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789_")
    result = input_str.replace(" ", "_").lower()
    return ''.join(c for c in result if c in allowed)

def fetch_existing_events(service):
    """
    Récupère les événements existants dans le calendrier, indexés par leur csv_id.
    """
    events_result = service.events().list(
        calendarId='primary',
        maxResults=1000,
        singleEvents=True
    ).execute()
    existing_events = {}
    for event in events_result.get('items', []):
        csv_id = event.get('extendedProperties', {}).get('private', {}).get('csv_id')
        if csv_id:
            existing_events[csv_id] = event
    return existing_events

def sync_events(service, df):
    """
    Synchronise les événements du CSV avec Google Calendar :
    création, mise à jour et suppression en batch.
    L'identifiant unique (csv_id) est généré à partir de Date, Start Time, Subject et Location.
    La couleur (colorId) est attribuée en fonction de la salle (avec une valeur par défaut de 5).
    """
    room_colors = {
        "Salle UT2J sans ordi": 11,
        "UT2J GS027": 6,
        "UT2J GS021": 5,
        "1003-Langue": 7,
        "Salle ENSAT sans ordi": 10,
        "703 (projet) ou alternance (entreprise)": 2,
        "UT2J GS028": 4,
    }
    existing_events = fetch_existing_events(service)
    events_to_create = []
    events_to_update = []
    events_to_delete = []
    for _, row in df.iterrows():
        # Génération stable du csv_id en combinant Date, Start Time, Subject et Location
        raw_id = f"{row['Date']}_{row['Start Time']}_{row['Subject']}_{row['Location']}"
        csv_id = sanitize_csv_id(raw_id)
        start_datetime = convert_to_datetime(row['Date'], row['Start Time'])
        end_datetime = convert_to_datetime(row['Date'], row['End Time'])
        location = row['Location']
        color_id = room_colors.get(location, 5)  # Couleur par défaut si la salle n'est pas référencée
        event_body = {
            "summary": row["Subject"],
            "location": location,
            "description": row["Description"],
            "start": {"dateTime": start_datetime.isoformat(), "timeZone": "Europe/Paris"},
            "end": {"dateTime": end_datetime.isoformat(), "timeZone": "Europe/Paris"},
            "extendedProperties": {"private": {"csv_id": csv_id}},
            "colorId": color_id,
        }
        if csv_id in existing_events:
            existing_event = existing_events[csv_id]
            differences = []
            if existing_event.get('summary', '') != row['Subject']:
                differences.append('summary')
            if existing_event.get('location', '') != location:
                differences.append('location')
            if existing_event.get('description', '') != row['Description']:
                differences.append('description')
            if existing_event.get('start', {}).get('dateTime', '') != start_datetime.isoformat():
                differences.append('start time')
            if existing_event.get('end', {}).get('dateTime', '') != end_datetime.isoformat():
                differences.append('end time')
            if differences:
                event_body['id'] = existing_event['id']
                events_to_update.append(event_body)
        else:
            events_to_create.append(event_body)
    existing_csv_ids = set(existing_events.keys())
    new_csv_ids = set(sanitize_csv_id(f"{row['Date']}_{row['Start Time']}_{row['Subject']}_{row['Location']}") for _, row in df.iterrows())
    obsolete_ids = existing_csv_ids - new_csv_ids
    for obsolete_id in obsolete_ids:
        events_to_delete.append(existing_events[obsolete_id]['id'])
    print(f"[Agenda] Création : {len(events_to_create)} | Mise à jour : {len(events_to_update)} | Suppression : {len(events_to_delete)}")
    batch = service.new_batch_http_request()
    for event in events_to_create:
        batch.add(service.events().insert(calendarId='primary', body=event))
    for event in events_to_update:
        batch.add(service.events().update(calendarId='primary', eventId=event['id'], body=event))
    for event_id in events_to_delete:
        batch.add(service.events().delete(calendarId='primary', eventId=event_id))
    batch.execute()

# =============================================================================
# PARTIE 5 : EXÉCUTION CONJOINTE AVEC THREADING
# =============================================================================

def run_modifications():
    """Lance la surveillance continue des modifications dans le fichier Excel."""
    surveiller_excel()

def run_agenda():
    """
    Exécute périodiquement la génération du CSV pour l'agenda et la synchronisation
    avec Google Calendar (toutes les 5 minutes).
    """
    while True:
        try:
            process_agenda()
        except Exception as e:
            print(f"Erreur dans le traitement de l'agenda : {e}")
        time.sleep(300)

def main():
    """
    Fonction principale qui démarre les threads de surveillance des modifications Excel
    et de génération/synchronisation de l'agenda.
    """
    t1 = threading.Thread(target=run_modifications, name="ModifThread")
    t2 = threading.Thread(target=run_agenda, name="AgendaThread")
    t1.start()
    t2.start()
    t1.join()
    t2.join()

if __name__ == "__main__":
    main()


